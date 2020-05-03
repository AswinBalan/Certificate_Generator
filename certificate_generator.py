import openpyxl

def generate(template, xlsx,domain,total):
    # print(total)
    # exit()
    rows = [] 
    fields= []
    wb_obj = openpyxl.load_workbook(xlsx) 
    sheet_obj = wb_obj.active 

    max_col = sheet_obj.max_column 
    max_row = sheet_obj.max_row
    print(max_col,max_row)
    
    # Will print a particular row value 
    for i in range(2, max_row + 1):
        tmp = []
        for j in range(1 ,max_col + 1):
            
            cell_obj = sheet_obj.cell(row = i, column = j)
            tmp.append(str(cell_obj.value))
        rows.append(tmp)
    print(rows)
    from drawing import draw
    draw(rows,template,domain,total,0)